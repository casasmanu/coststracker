import os
import logging
import datetime
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


def _ensure_excel(path: str):
    """Create Excel file if it does not exist"""
    if not os.path.exists(path):
        wb = Workbook()
        wb.save(path)
        wb.close()
        logger.info(f"Excel file created at {path}")


def _ensure_sheet(path: str, sheet_name: str, columns: list):
    """Create sheet with headers if it does not exist"""
    wb = load_workbook(path)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        wb.save(path)
        logger.info(f"Sheet '{sheet_name}' created")

    wb.close()


def updateExcel(path, sheetName, sheet_columns, sheet_data):
    """
    Append a row to a sheet.
    Automatically creates file and sheet if needed.
    """

    try:

        _ensure_excel(path)

        columns = sheet_columns.copy()

        if sheetName == "GastosVarios":
            columns = ["Date"] + columns

        _ensure_sheet(path, sheetName, columns)

        wb = load_workbook(path)
        ws = wb[sheetName]

        row = sheet_data.copy()

        if sheetName == "GastosVarios":
            row = [datetime.datetime.now()] + row

        ws.append(row)

        wb.save(path)
        wb.close()

        logger.info(f"Added row to {sheetName}: {row}")

        return True

    except Exception as e:
        logger.error(f"Error updating Excel: {e}")
        return False


def readData(path, sheetName):
    """Return sheet data formatted for Telegram"""

    try:

        if not os.path.exists(path):
            return "No Excel file found."

        wb = load_workbook(path)

        if sheetName not in wb.sheetnames:
            return f"No data for {sheetName}"

        ws = wb[sheetName]

        rows = list(ws.values)

        wb.close()

        if len(rows) <= 1:
            return "No data yet."

        header = rows[0]
        data_rows = rows[1:]

        output = f"📊 {sheetName} prices\n\n"

        for row in data_rows:
            row_text = " | ".join(str(x) for x in row)
            output += row_text + "\n"

        return output

    except Exception as e:
        logger.error(f"Error reading Excel: {e}")
        return "Error reading Excel."