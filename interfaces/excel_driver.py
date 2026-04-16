import os
import logging
import datetime
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

VARIABLE_EXPENSES_SHEET = "VariableExpenses"
LEGACY_VARIABLE_EXPENSES_SHEET = "GastosVarios"


def _resolve_variable_expenses_sheet(workbook, requested_sheet_name: str) -> str:
    """Map variable-expenses sheet to legacy name when needed."""
    if requested_sheet_name != VARIABLE_EXPENSES_SHEET:
        return requested_sheet_name

    if VARIABLE_EXPENSES_SHEET in workbook.sheetnames:
        return VARIABLE_EXPENSES_SHEET

    if LEGACY_VARIABLE_EXPENSES_SHEET in workbook.sheetnames:
        return LEGACY_VARIABLE_EXPENSES_SHEET

    return VARIABLE_EXPENSES_SHEET


def _ensure_excel(path: str):
    """Create Excel file if it does not exist"""
    if not os.path.exists(path):
        wb = Workbook()
        wb.save(path)
        wb.close()
        logger.info(f"Excel file created at {path}")


def _ensure_sheet(path: str, sheet_name: str, columns: list):
    """Create sheet with headers if it does not exist"""
    wb = load_workbook(path)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        wb.save(path)
        logger.info(f"Sheet '{sheet_name}' created")

    wb.close()


def update_excel(path, sheet_name, sheet_columns, sheet_data):
    """
    Append a row to a sheet.
    Automatically creates file and sheet if needed.
    """

    try:

        _ensure_excel(path)

        wb = load_workbook(path)
        resolved_sheet_name = _resolve_variable_expenses_sheet(wb, sheet_name)
        wb.close()

        columns = sheet_columns.copy()

        if resolved_sheet_name in (VARIABLE_EXPENSES_SHEET, LEGACY_VARIABLE_EXPENSES_SHEET):
            # Keep month explicitly in column F, leaving column E available.
            month_column = "Mes" if resolved_sheet_name == LEGACY_VARIABLE_EXPENSES_SHEET else "Month"
            columns = ["Date"] + columns + ["", month_column]

        _ensure_sheet(path, resolved_sheet_name, columns)

        wb = load_workbook(path)
        ws = wb[resolved_sheet_name]

        row = sheet_data.copy()

        if resolved_sheet_name in (VARIABLE_EXPENSES_SHEET, LEGACY_VARIABLE_EXPENSES_SHEET):
            today = datetime.date.today()
            month_names = {
                1: "JANUARY",
                2: "FEBRUARY",
                3: "MARCH",
                4: "APRIL",
                5: "MAY",
                6: "JUNE",
                7: "JULY",
                8: "AUGUST",
                9: "SEPTEMBER",
                10: "OCTOBER",
                11: "NOVEMBER",
                12: "DECEMBER",
            }
            month = month_names[today.month]
            # Row layout for VariableExpenses: A=date, B-D=user data, E=empty, F=month.
            row = [today] + row + ["", month]

        ws.append(row)

        wb.save(path)
        wb.close()

        logger.info(f"Added row to {resolved_sheet_name}: {row}")

        return True

    except Exception as e:
        logger.error(f"Error updating Excel: {e}")
        return False


def read_data(path, sheet_name):
    """Return sheet data formatted for Telegram"""

    try:

        if not os.path.exists(path):
            return "No Excel file found."

        wb = load_workbook(path)

        if sheet_name not in wb.sheetnames:
            return f"No data for {sheet_name}"

        ws = wb[sheet_name]

        rows = list(ws.values)

        wb.close()

        if len(rows) <= 1:
            return "No data yet."

        data_rows = rows[1:]

        output = f"{sheet_name} prices\n\n"

        for row in data_rows:
            row_text = " | ".join(str(x) for x in row)
            output += row_text + "\n"

        return output

    except Exception as e:
        logger.error(f"Error reading Excel: {e}")
        return "Error reading Excel."


def get_last_expenses(path: str, sheet_name: str = "VariableExpenses", n: int = 5):
    """Return the last n expense rows from the sheet as a list of dicts."""
    try:
        if not os.path.exists(path):
            return []

        wb = load_workbook(path)

        resolved_sheet_name = _resolve_variable_expenses_sheet(wb, sheet_name)

        if resolved_sheet_name not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[resolved_sheet_name]
        rows = list(ws.values)
        wb.close()

        if len(rows) <= 1:
            return []

        header = rows[0]
        data_rows = rows[1:]
        last_rows = data_rows[-n:]

        result = []
        for row in last_rows:
            result.append(dict(zip(header, row)))

        return result

    except Exception as e:
        logger.error(f"Error reading last expenses: {e}")
        return []


def _extract_year_month(date_val):
    if hasattr(date_val, "year") and hasattr(date_val, "month"):
        return date_val.year, date_val.month

    if isinstance(date_val, str):
        raw_date = date_val[:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                parsed = datetime.datetime.strptime(raw_date, fmt)
                return parsed.year, parsed.month
            except ValueError:
                continue

    return None, None


def get_expenses_for_month(path: str, year: int, month: int, sheet_name: str = "VariableExpenses"):
    """Return all expenses for a given month and year as a list of dicts."""
    try:
        if not os.path.exists(path):
            return []

        wb = load_workbook(path)
        resolved_sheet_name = _resolve_variable_expenses_sheet(wb, sheet_name)

        if resolved_sheet_name not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[resolved_sheet_name]
        rows = list(ws.values)
        wb.close()

        if len(rows) <= 1:
            return []

        header = rows[0]
        normalized_header = [str(c).lower() if c else "" for c in header]
        date_col = next((i for i, c in enumerate(normalized_header) if c == "date"), None)

        if date_col is None:
            return []

        month_rows = []
        for row in rows[1:]:
            date_val = row[date_col] if len(row) > date_col else None
            row_year, row_month = _extract_year_month(date_val)
            if row_year == year and row_month == month:
                month_rows.append(dict(zip(header, row)))

        return month_rows

    except Exception as e:
        logger.error(f"Error reading monthly expenses: {e}")
        return []


def get_monthly_category_breakdown(path: str, year: int, month: int, sheet_name: str = "VariableExpenses"):
    """Return category totals for a given month and year sorted by highest amount."""
    try:
        expenses = get_expenses_for_month(path=path, year=year, month=month, sheet_name=sheet_name)
        if not expenses:
            return []

        keys = [str(k).lower() if k else "" for k in expenses[0].keys()]
        original_keys = list(expenses[0].keys())

        amount_key = next(
            (original_keys[i] for i, k in enumerate(keys) if k in ("amount", "cuantity")),
            None,
        )
        category_key = next(
            (original_keys[i] for i, k in enumerate(keys) if k in ("extra", "category", "comments")),
            None,
        )
        description_key = next(
            (original_keys[i] for i, k in enumerate(keys) if k == "description"),
            None,
        )

        if amount_key is None:
            return []

        totals = {}
        for expense in expenses:
            raw_amount = expense.get(amount_key)
            try:
                amount = float(raw_amount)
            except (TypeError, ValueError):
                continue

            category = expense.get(category_key) if category_key else None
            if category in (None, "") and description_key:
                category = expense.get(description_key)
            category = str(category).strip() if category not in (None, "") else "Other"

            totals[category] = totals.get(category, 0.0) + amount

        return sorted(totals.items(), key=lambda item: item[1], reverse=True)

    except Exception as e:
        logger.error(f"Error building monthly category breakdown: {e}")
        return []


def get_monthly_total(path: str, month: int, sheet_name: str = "VariableExpenses") -> float:
    """Return total expenses for a given month number."""
    try:
        if not os.path.exists(path):
            return 0.0

        wb = load_workbook(path)
        resolved_sheet_name = _resolve_variable_expenses_sheet(wb, sheet_name)

        if resolved_sheet_name not in wb.sheetnames:
            wb.close()
            return 0.0

        ws = wb[resolved_sheet_name]
        rows = list(ws.values)
        wb.close()

        if len(rows) <= 1:
            return 0.0

        header = [str(c).lower() if c else "" for c in rows[0]]

        amount_col = next((i for i, c in enumerate(header) if c in ("amount", "cuantity")), None)
        date_col = next((i for i, c in enumerate(header) if c == "date"), None)

        if amount_col is None or date_col is None:
            return 0.0

        total = 0.0
        for row in rows[1:]:
            date_val = row[date_col] if len(row) > date_col else None
            row_month = None
            if hasattr(date_val, "month"):
                row_month = date_val.month
            elif isinstance(date_val, str):
                try:
                    row_month = datetime.datetime.strptime(date_val[:10], "%Y-%m-%d").month
                except ValueError:
                    pass
            if row_month == month and len(row) > amount_col:
                try:
                    total += float(row[amount_col])
                except (ValueError, TypeError):
                    pass

        return total

    except Exception as e:
        logger.error(f"Error getting monthly total: {e}")
        return 0.0


def delete_last_expense(path: str, sheet_name: str = "VariableExpenses") -> dict | None:
    """Delete the last expense row and return it as a dict, or None if nothing was deleted."""
    try:
        if not os.path.exists(path):
            return None

        wb = load_workbook(path)
        resolved_sheet_name = _resolve_variable_expenses_sheet(wb, sheet_name)

        if resolved_sheet_name not in wb.sheetnames:
            wb.close()
            return None

        ws = wb[resolved_sheet_name]

        if ws.max_row <= 1:
            wb.close()
            return None

        header = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        last_row_values = [ws.cell(ws.max_row, col).value for col in range(1, ws.max_column + 1)]
        deleted = dict(zip(header, last_row_values))

        ws.delete_rows(ws.max_row)
        wb.save(path)
        wb.close()

        return deleted

    except Exception as e:
        logger.error(f"Error deleting last expense: {e}")
        return None
